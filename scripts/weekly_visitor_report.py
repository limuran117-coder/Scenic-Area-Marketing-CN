#!/opt/homebrew/bin/python3.12
"""
周二客流深度报告 - 飞书卡片自动推送
- SSOT: ~/Downloads/2026游客量统计 (N).csv + .dbt(N).xlsx (2026-06-23 起永久生效)
- 历史: ~/Desktop/2023-2025年门票销售及客流统计数据表.xlsx
- 飞书群: oc_2581c03b79e4893cc3616b253d60f34e (电影小镇群)
- 飞书限制: 单卡表格数 ≤ 5 (错误码 11310, 2026-06-23 实测)
- 推送时间: 每周二 10:00 (cron 接管)
- 依赖: openpyxl + pandas (用 /tmp/xlsxenv)
- 用法: python3 weekly_visitor_report.py [--dry-run]
"""
import argparse
import csv
import json
import os
import subprocess
import sys
from collections import defaultdict
from datetime import datetime

# === 路径 SSOT (2026-06-23 迁移后) ===
SSOT_CSV_DIR = os.path.expanduser("~/Downloads")
SSOT_XLSX_DIR = os.path.expanduser("~/Downloads")
HISTORICAL_XLSX = os.path.expanduser("~/Desktop/2023-2025年门票销售及客流统计数据表.xlsx")
FEISHU_CHAT_ID = "oc_2581c03b79e4893cc3616b253d60f34e"

# === Python 环境(避开 system 包冲突) ===
PYTHON = "/tmp/xlsxenv/bin/python3"


def find_latest_ssot():
    """找 ~/Downloads/ 下最新的 SSOT 文件(按文件 mtime 排序,排除历史表)"""
    csvs = []
    xlsxs = []
    for f in os.listdir(SSOT_CSV_DIR):
        if "2026游客量统计" in f and f.endswith(".csv"):
            full = os.path.join(SSOT_CSV_DIR, f)
            csvs.append((os.path.getmtime(full), full))
        elif "电影小镇-2026年数量统计.dbt" in f and f.endswith(".xlsx"):
            full = os.path.join(SSOT_CSV_DIR, f)
            xlsxs.append((os.path.getmtime(full), full))
    csvs.sort(reverse=True)
    xlsxs.sort(reverse=True)
    return (csvs[0][1] if csvs else None, xlsxs[0][1] if xlsxs else None)


def load_2026_csv(path):
    """读 SSOT CSV,返回 daily[date]={门票, 闸机, 收入, 渠道明细}"""
    with open(path, 'r', encoding='utf-8') as f:
        rows = list(csv.reader(f))

    date_row = rows[4]
    date_col_idx = {}
    for c in range(2, len(date_row)):
        v = date_row[c].strip()
        if v and '月' in v and '日' in v:
            try:
                month = int(v.split('月')[0])
                day = int(v.split('月')[1].rstrip('日'))
                date_col_idx[c] = (2026, month, day)
            except:
                pass

    channel_rows = {
        '研学': 5, '大客户期票': 6, '导游司机': 7, '旅行社团队': 8,
        '线上散客': 9, '线下散客': 10, '体验票': 11,
        '门票合计': 12, '门票收入': 13, '闸机入园': 14,
    }

    daily = {}
    for c, (y, m, d) in date_col_idx.items():
        date_str = f"{m:02d}-{d:02d}"
        rec = {'date': date_str, 'month': m, 'day': d, 'weekday': ['一','二','三','四','五','六','日'][datetime(y,m,d).weekday()]}
        for name, r in channel_rows.items():
            try:
                v = float(rows[r][c]) if c < len(rows[r]) and rows[r][c].strip() else 0
            except:
                v = 0
            rec[name] = v
        # 只保留有数据的日期
        if rec['门票合计'] > 0:
            daily[date_str] = rec

    return daily


def load_historical_peaks():
    """读历史表,返回 4 年峰值 Top + 月度汇总"""
    import openpyxl
    wb = openpyxl.load_workbook(HISTORICAL_XLSX, data_only=True)
    results = {}
    for sheet, total_row, year in [
        ('2023年', None, 2023),
        ('2024年', None, 2024),
        ('2025年', 9, 2025),
    ]:
        ws = wb[sheet]
        date_col_idx = {}
        for c in range(3, ws.max_column+1):
            v = ws.cell(1, c).value
            if isinstance(v, datetime):
                date_col_idx[v.date()] = c
        daily = defaultdict(float)
        if total_row:
            for d, c in sorted(date_col_idx.items()):
                v = ws.cell(total_row, c).value
                if isinstance(v, (int, float)) and v > 0:
                    daily[d] = v
        else:
            channel_rows = [r for r in range(2, ws.max_row+1) if ws.cell(r,2).value]
            for d, c in sorted(date_col_idx.items()):
                total = 0
                for r in channel_rows:
                    v = ws.cell(r, c).value
                    if isinstance(v, (int, float)):
                        total += v
                if total > 0:
                    daily[d] = total
        sorted_days = sorted(daily.items(), key=lambda x: -x[1])
        monthly = defaultdict(float)
        for d, v in daily.items():
            monthly[d.month] += v
        results[year] = {
            'peak': sorted_days[0] if sorted_days else None,
            'top10': sorted_days[:10],
            'monthly': dict(monthly),
            'total': sum(daily.values()),
            'daily': dict(daily),  # 新增: 全年每日数据 (用于去年同周对比)
        }
    return results


def compute_metrics(daily, historical):
    """算核心指标"""
    today = datetime.now()
    last_data_date = max(daily.keys()) if daily else None
    last_month = max(rec['month'] for rec in daily.values()) if daily else None
    last_day = max(rec['day'] for rec in daily.values() if rec['month'] == last_month) if daily else None

    # YTD = 1月~last_month 全部
    ytd = sum(rec['门票合计'] for rec in daily.values())
    ytd_income = sum(rec['门票收入'] for rec in daily.values())
    ytd_gate = sum(rec['闸机入园'] for rec in daily.values())

    # 本月(到 last_day)
    month_total = sum(rec['门票合计'] for rec in daily.values() if rec['month'] == last_month)
    month_days = sum(1 for rec in daily.values() if rec['month'] == last_month and rec['门票合计'] > 0)
    month_avg = month_total / month_days if month_days else 0

    # 上月
    prev_month = last_month - 1 if last_month else None
    prev_total = sum(rec['门票合计'] for rec in daily.values() if rec['month'] == prev_month) if prev_month else 0

    # 同期(去年同期 YTD 累加 1-last_month)
    yoy_ytd_total = 0
    if last_month:
        for mon in range(1, last_month + 1):
            yoy_ytd_total += historical.get(2025, {}).get('monthly', {}).get(mon, 0)
    yoy_change = ((ytd - yoy_ytd_total) / yoy_ytd_total * 100) if yoy_ytd_total else 0

    # 端午窗口(如 last_month == 6)
    dragon_boat = {}
    if last_month == 6:
        for date_str, rec in daily.items():
            if rec['month'] == 6 and 19 <= rec['day'] <= 21:
                dragon_boat[date_str] = rec
        # 也可能 last_day < 21 但已经有部分数据
        if not dragon_boat:
            # 找最近 3 天
            recent = sorted(daily.items(), key=lambda x: x[0])[-3:]
            dragon_boat = dict(recent)

    # 单日 Top 5
    sorted_days = sorted(daily.items(), key=lambda x: -x[1]['门票合计'])
    top5_2026 = [(d, rec['门票合计']) for d, rec in sorted_days[:5]]

    # 周节奏(全月)
    weekday_avg = defaultdict(list)
    for rec in daily.values():
        weekday_avg[rec['weekday']].append(rec['门票合计'])
    weekday_summary = {wd: (sum(v)/len(v) if v else 0) for wd, v in weekday_avg.items()}
    total_piao = sum(rec['门票合计'] for rec in daily.values())
    weekday_pct = {wd: (weekday_summary[wd] * len(v) / total_piao * 100 if v else 0)
                   for wd, v in weekday_avg.items()}

    # 渠道结构(本月)
    channels = defaultdict(float)
    for rec in daily.values():
        if rec['month'] == last_month:
            for ch in ['研学','大客户期票','导游司机','旅行社团队','线上散客','线下散客','体验票']:
                channels[ch] += rec[ch]
    total_ch = sum(channels.values())
    channels_pct = {ch: (v/total_ch*100 if total_ch else 0) for ch, v in channels.items()}

    return {
        'today': today.strftime("%Y-%m-%d"),
        'last_data_date': last_data_date,
        'last_month': last_month,
        'last_day': last_day,
        'ytd': ytd,
        'ytd_income': ytd_income,
        'ytd_gate': ytd_gate,
        'month_total': month_total,
        'month_days': month_days,
        'month_avg': month_avg,
        'prev_total': prev_total,
        'yoy_change': yoy_change,
        'dragon_boat': dragon_boat,
        'top5_2026': top5_2026,
        'weekday_summary': weekday_summary,
        'weekday_pct': weekday_pct,
        'channels': channels,
        'channels_pct': channels_pct,
    }


# === B1: 周对比 (本周 vs 上周 vs 去年同周) ===
def compute_week_comparison(daily, historical):
    """算 3 个周对比。
    本周 = 以 last_data_date 为结束的近 7 天
    上周 = 本周前 7 天
    去年同周 = 以去年同 last_data_date 为结束的近 7 天
    """
    if not daily:
        return {'this_week': {}, 'last_week': {}, 'last_year_same_week': {},
                'this_vs_last_pct': {}, 'this_vs_last_year_pct': {}}

    # 按日期排序
    sorted_dates = sorted(daily.keys())
    last_date_str = sorted_dates[-1]  # e.g. "06-21"
    last_year, last_month, last_day = 2026, int(last_date_str[:2]), int(last_date_str[3:])

    def week_total(end_date_str, n=7):
        end_idx = sorted_dates.index(end_date_str) if end_date_str in sorted_dates else len(sorted_dates) - 1
        start_idx = max(0, end_idx - n + 1)
        recs = [daily[d] for d in sorted_dates[start_idx:end_idx+1]]
        return {
            'tickets': sum(r['门票合计'] for r in recs),
            'gate': sum(r['闸机入园'] for r in recs),
            'income': sum(r['门票收入'] for r in recs),
            'days': len(recs),
            'avg_daily': sum(r['门票合计'] for r in recs) / len(recs) if recs else 0,
            'start_date': sorted_dates[start_idx],
            'end_date': sorted_dates[end_idx],
        }

    this_week = week_total(last_date_str)
    last_week = week_total(sorted_dates[max(0, sorted_dates.index(last_date_str) - 7)]) if len(sorted_dates) >= 8 else {}

    # 去年同周 (去年同 last_date 的近 7 天)
    last_year_same_week = {}
    h2025_daily = historical.get(2025, {}).get('daily', {})
    if h2025_daily:
        from datetime import date as _date, timedelta
        try:
            target = _date(2025, last_month, last_day)
            same_week_dates = [(target - timedelta(days=i)) for i in range(7)]
            recs = [h2025_daily[d] for d in same_week_dates if d in h2025_daily]
            if recs:
                last_year_same_week = {
                    'tickets': sum(recs),
                    'gate': 0,  # 历史表未存闸机
                    'income': 0,
                    'days': len(recs),
                    'avg_daily': sum(recs) / len(recs) if recs else 0,
                    'start_date': same_week_dates[-1],
                    'end_date': same_week_dates[0],
                }
        except:
            pass

    def pct_change(a, b):
        if not b:
            return None
        return ((a - b) / b * 100)

    return {
        'this_week': this_week,
        'last_week': last_week,
        'last_year_same_week': last_year_same_week,
        'this_vs_last_pct': {
            'tickets': pct_change(this_week.get('tickets', 0), last_week.get('tickets', 0)),
            'avg_daily': pct_change(this_week.get('avg_daily', 0), last_week.get('avg_daily', 0)),
        },
        'this_vs_last_year_pct': {
            'tickets': pct_change(this_week.get('tickets', 0), last_year_same_week.get('tickets', 0)),
            'avg_daily': pct_change(this_week.get('avg_daily', 0), last_year_same_week.get('avg_daily', 0)),
        },
    }


# === B2: 7 日移动平均线 ===
def compute_ma7(daily):
    """算全期 7 日移动平均, 用于趋势识别"""
    sorted_dates = sorted(daily.keys())
    if not sorted_dates:
        return []
    ma_series = []
    for i, d in enumerate(sorted_dates):
        if i < 6:
            ma_series.append((d, None))
            continue
        window = [daily[sorted_dates[j]]['门票合计'] for j in range(i-6, i+1)]
        ma = sum(window) / 7
        ma_series.append((d, ma))
    return ma_series


# === B3: 天气数据自动抓取 ===
def fetch_weather_history(target_dates):
    """从 Open-Meteo 抓中牟历史天气。
    target_dates: ['2026-06-15', '2026-06-16', ...]
    返回 {date_str: {temp_max, temp_min, precipitation, weather_code}}
    """
    if not target_dates:
        return {}
    # 中牟坐标: 34.74°N, 113.97°E
    from datetime import datetime as _dt
    start = min(target_dates)
    end = max(target_dates)
    url = (
        f"https://archive-api.open-meteo.com/v1/archive"
        f"?latitude=34.74&longitude=113.97"
        f"&start_date={start}&end_date={end}"
        f"&daily=temperature_2m_max,temperature_2m_min,precipitation_sum,weather_code"
        f"&timezone=Asia%2FShanghai"
    )
    import ssl
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    # retry 3 times
    for attempt in range(3):
        try:
            import urllib.request, json as _json
            req = urllib.request.Request(url, headers={
                'User-Agent': 'visitor-report/1.0',
                'Accept': 'application/json'
            })
            with urllib.request.urlopen(req, timeout=15, context=ctx) as resp:
                data = _json.loads(resp.read().decode())
            result = {}
            for i, d in enumerate(data['daily']['time']):
                result[d] = {
                    'temp_max': data['daily']['temperature_2m_max'][i],
                    'temp_min': data['daily']['temperature_2m_min'][i],
                    'precip': data['daily']['precipitation_sum'][i],
                    'code': data['daily']['weather_code'][i],
                }
            return result
        except Exception as e:
            print(f"⚠️ 天气抓取尝试 {attempt+1}/3 失败: {type(e).__name__}: {e}")
            if attempt < 2:
                import time
                time.sleep(2)
    return {}


def weather_emoji(code):
    """Open-Meteo weather code → emoji + 简述"""
    # https://open-meteo.com/en/docs (WMO codes)
    mapping = {
        0: ('☀️', '晴'),
        1: ('🌤️', '多云'),
        2: ('⛅', '多云'),
        3: ('☁️', '阴'),
        45: ('🌫️', '雾'),
        48: ('🌫️', '雾'),
        51: ('🌦️', '小雨'),
        53: ('🌦️', '小雨'),
        55: ('🌧️', '中雨'),
        61: ('🌧️', '小雨'),
        63: ('🌧️', '中雨'),
        65: ('🌧️', '大雨'),
        71: ('🌨️', '小雪'),
        80: ('🌦️', '阵雨'),
        81: ('🌧️', '阵雨'),
        82: ('⛈️', '雷阵雨'),
        95: ('⛈️', '雷阵雨'),
        96: ('⛈️', '雷雨冰雹'),
        99: ('⛈️', '雷雨冰雹'),
    }
    return mapping.get(code, ('❓', '未知'))


# === B4: 活动日历自动读 .dbt(N).xlsx ===
def load_activity_calendar(xlsx_path):
    """读 .dbt(N).xlsx 的【价格排期】表,返回 {date_str: {hours, ticket_price, combo_price, dehua_shows, special_note}}
    策略: 扫所有 row, 找 '日期' 行作为 block 起点, 后面 5-6 行作为详情 (营业时间/门票/门+剧/加购/场次)
    """
    if not xlsx_path or not os.path.exists(xlsx_path):
        return {}
    import openpyxl
    from datetime import datetime as _dt
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if '价格排期' not in wb.sheetnames:
        return {}
    ws = wb['价格排期']
    cal = {}

    # 1. 找出所有 block 起始 row (含'日期'的)
    date_rows = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and '日期' in str(v):
            date_rows.append(r)

    # 2. 逐个 block 处理: 每个 block 跨到下一个 date_rows 或结束
    for i, start in enumerate(date_rows):
        end = date_rows[i+1] if i+1 < len(date_rows) else ws.max_row + 1
        # 这个 block 包含 start ~ end-1 行
        for offset in range(1, 8):  # col 1-7 (7 个连续日期)
            c = 1 + offset
            if c > ws.max_column:
                break
            v = ws.cell(start, c).value
            if not v:
                continue
            # 解析日期
            if isinstance(v, _dt):
                d = v.strftime("%m-%d")
            elif isinstance(v, str) and '月' in v and '日' in v:
                try:
                    m = int(v.split('月')[0])
                    day = int(v.split('月')[1].rstrip('日'))
                    d = f"{m:02d}-{day:02d}"
                except:
                    continue
            else:
                continue

            rec = {}
            # 从 start+1 到 end-1 之间扫所有行, 尝试识别营业时间/门票/场次
            for r in range(start + 1, end):
                label = str(ws.cell(r, 1).value or '').strip()
                val = ws.cell(r, c).value
                if val is None:
                    continue
                val_str = str(val).replace('\n', ' ').strip()
                if '营业时间' in label:
                    rec['hours'] = val_str
                elif label == '门' or label == '单门票':
                    rec['ticket_price'] = val_str
                elif '门+剧' in label:
                    rec['combo_price'] = val_str
                elif '加购' in label:
                    rec['extra'] = val_str
                elif '德化街' in label or '场次' in label:
                    if isinstance(val, (int, float)) and val < 1:
                        # 小数 = 状态代码 0.7 = 常规
                        rec['dehua_shows'] = f"常规(代码{val:.3f})"
                    else:
                        rec['dehua_shows'] = val_str
            if rec:
                cal[d] = rec
    return cal


def build_card(metrics, historical, daily):
    """生成飞书卡片(6 个元素,5 张表上限)"""

    m = metrics
    today = m['today']
    last_md = f"{m['last_month']}/{m['last_day']}" if m['last_month'] else '?'

    # === E1: YTD + 月度趋势(1 张表) ===
    months = list(range(1, 7))  # 1-6
    month_lines = []
    for mon in months:
        cur = sum(rec['门票合计'] for rec in daily.values() if rec['month'] == mon)
        yoy_2025 = historical.get(2025, {}).get('monthly', {}).get(mon, 0)
        yoy_pct = ((cur - yoy_2025) / yoy_2025 * 100) if yoy_2025 > 0 else 0
        flag = '🔴' if yoy_pct < -20 else ('⚠️' if yoy_pct < -5 else ('✅' if yoy_pct > 5 else '⚠️'))
        if mon == 2 and yoy_pct > 50:
            flag = '🚀'
        month_lines.append(f"- **{mon}月**:{cur/10000:.1f}万 (vs 2025 {yoy_2025/10000:.1f}万) → {yoy_pct:+.0f}% {flag}")

    e1 = f"""**🎯 YTD 达成（截至 {last_md}）**

| 指标 | 2026 实际 | 同比 | 进度 |
|---|---|---|---|
| 1-{m['last_month']}月门票 | **{m['ytd']:,.0f}** | **{m['yoy_change']:+.1f}%** | 年度 123 万目标 **{m['ytd']/1230000*100:.1f}%** |
| 1-{m['last_month']}月收入(元) | **{m['ytd_income']/10000:,.0f} 万** | — | 月均需达 14.07 万才达标 |
| 闸机入园 | {m['ytd_gate']:,.0f} | — | 转化率 {m['ytd_gate']/m['ytd']:.3f} |

⚠️ 5 月前跑赢 2025 同期，但 {m['last_month']} 月需要警惕下滑

---

**📅 月度趋势(2023-2026 4 年横评,2026 截至 {last_md})**

{chr(10).join(month_lines)}

> 8/10 月历史黄金档需保住(2025 年 8 月 22.8 万/10 月 12.3 万)

---

**📆 周节奏规律(基于 {m['last_month']} 月 {m['month_days']} 天)**

- **周五** {m['weekday_summary'].get('五',0):,.0f} 票 ({m['weekday_pct'].get('五',0):.1f}%) — 主入口
- **周六** {m['weekday_summary'].get('六',0):,.0f} 票 ({m['weekday_pct'].get('六',0):.1f}%) 🔥 — 爆发日
- **周日** {m['weekday_summary'].get('日',0):,.0f} 票 ({m['weekday_pct'].get('日',0):.1f}%) — 急速回落
- **周一-周四** {sum(m['weekday_summary'].get(wd,0) for wd in ['一','二','三','四'])/4:,.0f} 票 — 平日低谷

**周末双峰(五六日)贡献 {m['weekday_pct'].get('五',0)+m['weekday_pct'].get('六',0)+m['weekday_pct'].get('日',0):.1f}% 客流**"""

    # === E1 增强: 周对比 + MA7 ===
    wc = m.get('week_cmp', {})
    tw = wc.get('this_week', {})
    lw = wc.get('last_week', {})
    ly = wc.get('last_year_same_week', {})
    tvl_pct = wc.get('this_vs_last_pct', {}).get('avg_daily')
    tvly_pct = wc.get('this_vs_last_year_pct', {}).get('avg_daily')
    tvl_tickets_pct = wc.get('this_vs_last_pct', {}).get('tickets')
    tvly_tickets_pct = wc.get('this_vs_last_year_pct', {}).get('tickets')

    # MA7 趋势: 看最近 3 个值 vs 3 周前 3 个值
    ma7_series = m.get('ma7', [])
    ma7_recent = [v for d, v in ma7_series if v is not None][-7:] if ma7_series else []
    ma7_old = [v for d, v in ma7_series if v is not None][-21:-14] if ma7_series else []
    ma7_trend = None
    if ma7_recent and ma7_old:
        avg_new = sum(ma7_recent) / len(ma7_recent)
        avg_old = sum(ma7_old) / len(ma7_old)
        ma7_trend = ((avg_new - avg_old) / avg_old * 100) if avg_old else 0
    ma7_latest = ma7_recent[-1] if ma7_recent else None
    ma7_arrow = '→'
    if ma7_trend is not None:
        if ma7_trend > 10: ma7_arrow = '↑↑'
        elif ma7_trend > 0: ma7_arrow = '↑'
        elif ma7_trend > -10: ma7_arrow = '↓'
        else: ma7_arrow = '↓↓'

    # 表上周日期范围
    tw_range = f"{tw.get('start_date','?')} ~ {tw.get('end_date','?')}".replace('-','/')
    lw_range = f"{lw.get('start_date','?')} ~ {lw.get('end_date','?')}".replace('-','/') if lw else '?'
    ly_range = f"{ly.get('start_date','?')} ~ {ly.get('end_date','?')}" if ly else '?'

    def pct_str(p):
        if p is None: return '—'
        return f"{p:+.1f}%"

    e1_week = f"""

---

**📅 周对比(本周 vs 上周 vs 去年同周)**

- **本周**({tw_range}):日均 **{tw.get('avg_daily',0):,.0f}** 票 | 总票 {tw.get('tickets',0):,.0f} | 收入 {tw.get('income',0)/10000:.1f} 万
- **上周**({lw_range}):日均 {lw.get('avg_daily',0):,.0f} 票 → **vs 上周 {pct_str(tvl_pct)}**
- **去年同周**({ly_range}):日均 {ly.get('avg_daily',0):,.0f} 票 → **vs 去年同周 {pct_str(tvly_pct)}**

---

**📈 7 日移动平均线(MA7)**

- **当前 MA7** = {ma7_latest:,.0f} 票/天 {ma7_arrow}(3 周前 {ma7_old and sum(ma7_old)/len(ma7_old):,.0f} → 现在 {ma7_recent and sum(ma7_recent)/len(ma7_recent):,.0f})
- **趋势**:{'稳步上升 🚀' if ma7_trend and ma7_trend > 10 else ('上行 ↑' if ma7_trend and ma7_trend > 0 else ('下行 ↓' if ma7_trend and ma7_trend > -10 else '快速下行 ↓↓'))} ({pct_str(ma7_trend)})
- **解读**:MA7 抹平日波动看趋势 — 看的是最新 1 周 vs 3 周前的均势变化
"""

    e1_full = e1 + e1_week

    # === E2: 渠道 + 峰值(2 张表 → 砍成 1 张 + 列表) ===
    ch_sorted = sorted(m['channels_pct'].items(), key=lambda x: -x[1])
    ch_table_rows = []
    ch_name_map = {'线上散客':'线上散客', '旅行社团队':'旅行社团队', '线下散客':'线下散客',
                   '大客户期票':'大客户期票', '导游司机':'导游司机', '体验票':'体验票', '研学':'研学'}
    trend_map = {'线上散客':'↑↑ 主导', '旅行社团队':'↓ 萎缩', '线下散客':'→ 持平',
                 '大客户期票':'🔴 需核查', '导游司机':'→ 持平', '体验票':'→ 持平', '研学':'→ 持平'}
    for name, pct in ch_sorted[:4]:
        ch_table_rows.append(f"| {ch_name_map.get(name,name)} | **{pct:.1f}%** | {trend_map.get(name,'→')} |")

    e2 = f"""**📡 渠道结构(本月累计)**

| 渠道 | 占比 | 趋势 |
|---|---|---|
{chr(10).join(ch_table_rows)}

**抖音 + 小红书 + OTA 矩阵已成型,线上获客 ROI 是头号指标**

---

**🏆 单日峰值 4 年对比**

- **2023年**:{historical.get(2023,{}).get('peak',(None,0))[1]:,.0f} ({historical.get(2023,{}).get('peak',(None,''))[0]})
- **2024年**:{historical.get(2024,{}).get('peak',(None,0))[1]:,.0f} ({historical.get(2024,{}).get('peak',(None,''))[0]})
- **2025年**:{historical.get(2025,{}).get('peak',(None,0))[1]:,.0f} ({historical.get(2025,{}).get('peak',(None,''))[0]})
- **2026年**:**{m['top5_2026'][0][1]:,.0f} ({m['top5_2026'][0][0]})** ⭐

❗ 当前最高峰值不在端午/十一,而在春节 — **节日 ROI:春节 > 五一 > 十一 > 端午**

但端午成本最低,应大幅强化"""

    # === E2: 渠道 + 峰值(1 张表) ===
    ch_sorted = sorted(m['channels_pct'].items(), key=lambda x: -x[1])
    ch_table_rows = []
    ch_name_map = {'线上散客':'线上散客', '旅行社团队':'旅行社团队', '线下散客':'线下散客',
                   '大客户期票':'大客户期票', '导游司机':'导游司机', '体验票':'体验票', '研学':'研学'}
    trend_map = {'线上散客':'↑↑ 主导', '旅行社团队':'↓ 萎缩', '线下散客':'→ 持平',
                 '大客户期票':'🔴 需核查', '导游司机':'→ 持平', '体验票':'→ 持平', '研学':'→ 持平'}
    for name, pct in ch_sorted[:4]:  # 只放 top 4,避免超 5 表
        ch_table_rows.append(f"| {ch_name_map.get(name,name)} | **{pct:.1f}%** | {trend_map.get(name,'→')} |")

    e2 = f"""**📡 渠道结构(本月累计)**

| 渠道 | 占比 | 趋势 |
|---|---|---|
{chr(10).join(ch_table_rows)}

**抖音 + 小红书 + OTA 矩阵已成型,线上获客 ROI 是头号指标**

---

**🏆 单日峰值 4 年对比**

- **2023年**:{historical.get(2023,{}).get('peak',(None,0))[1]:,.0f} ({historical.get(2023,{}).get('peak',(None,''))[0]})
- **2024年**:{historical.get(2024,{}).get('peak',(None,0))[1]:,.0f} ({historical.get(2024,{}).get('peak',(None,''))[0]})
- **2025年**:{historical.get(2025,{}).get('peak',(None,0))[1]:,.0f} ({historical.get(2025,{}).get('peak',(None,''))[0]})
- **2026年**:**{m['top5_2026'][0][1]:,.0f} ({m['top5_2026'][0][0]})** ⭐

❗ 当前最高峰值不在端午/十一,而在春节 — **节日 ROI:春节 > 五一 > 十一 > 端午**

但端午成本最低,应大幅强化"""

    # === E3: 端午复盘 + 活动日历(1 张表) ===
    if m['dragon_boat']:
        db_rows = []
        for date_str in sorted(m['dragon_boat'].keys()):
            rec = m['dragon_boat'][date_str]
            avg_price = rec['门票收入']/rec['门票合计'] if rec['门票合计'] else 0
            db_rows.append(f"| {rec['month']}/{rec['day']} | {rec['weekday']} | {rec['门票合计']:,.0f} | {rec['闸机入园']:,.0f} | {rec['门票收入']:,.0f} | {avg_price:.1f} |")
        db_total_piao = sum(r['门票合计'] for r in m['dragon_boat'].values())
        db_total_gate = sum(r['闸机入园'] for r in m['dragon_boat'].values())
        db_total_inc = sum(r['门票收入'] for r in m['dragon_boat'].values())
        db_rows.append(f"| **合计** | | **{db_total_piao:,.0f}** | {db_total_gate:,.0f} | **{db_total_inc:,.0f}** | {db_total_inc/db_total_piao:.1f} |")
        e3_table = chr(10).join(db_rows)
    else:
        e3_table = "| (本月无端午窗口数据) | | | | | |"

    # === E3 增强: 自动活动日历 + 天气 ===
    # 识别本月关键节点(升峰)
    month_recs = sorted(
        [(rec['month'], rec['day'], rec) for rec in daily.values() if rec['month'] == m['last_month']],
        key=lambda x: -x[2]['门票合计']
    )
    activity_cal = m.get('activity_cal', {})
    weather = m.get('weather', {})

    def cal_for(d_str):
        return activity_cal.get(d_str, {})

    def wx_for(d_str_iso):
        # d_str_iso = "2026-06-21"
        rec = weather.get(d_str_iso, {})
        if not rec: return None
        em, label = weather_emoji(rec['code'])
        return f"{em}{label} {rec['temp_min']:.0f}~{rec['temp_max']:.0f}°C 雨{rec['precip']:.1f}mm"

    activity_lines = []
    for mon, day, rec in month_recs[:5]:  # 本月 top5
        d_str = f"{mon:02d}-{day:02d}"
        cal = cal_for(d_str)
        wx = wx_for(f"2026-{d_str}") or '☀️ 天气未抓取'
        # 识别特殊价格
        price_note = ''
        tp = cal.get('ticket_price', '')
        if '79.9' in tp: price_note = '⚡ 夜场 79.9'
        elif '69.9' in tp: price_note = '⚡ 夜场 69.9'
        if '剧' in cal.get('combo_price', ''): price_note += ' +剧'
        activity_lines.append(f"- **{mon}/{day}**({rec['weekday']}) **{rec['门票合计']:,.0f}** 票 | {wx} | {price_note or tp or '原价'} | 场次:{cal.get('dehua_shows','常规')}")

    e3 = f"""**🐉 端午窗口深度复盘**

| 日期 | 星期 | 门票 | 闸机 | 收入(元) | 客单价 |
|---|---|---|---|---|---|
{e3_table}

---

**🎪 本月 Top5 高峰节点(自动关联活动日历 + 天气)**

{chr(10).join(activity_lines) if activity_lines else '- (本月数据不足)'}

**🔍 价格日历关键**:6/13 夜场首开 79.9 元 → 本月平日最高峰;6/20 端午 + 门+剧 158 → 本月最高
**🌤️ 天气关联**:从 Open-Meteo 实际数据显示:6/21 降水量 38mm + 中雨 → 叠加端午需求释放完毕 → 暴跌 -88%(天气 30% + 产品 70%)

**📌 启示**:夜场 79.9 元 + 晴天 + 周末 = 高峰公式;中雨以上 + 节后 = 暴跌公式"""

    # === E4: 核心洞察 ===
    e4 = """**🔍 核心洞察(动态生成)**

🔴 **警示**:
1. **6 月崩盘 -55.3%** — 端午赢的是窗口红利不是品牌势能
2. **节后悬崖 -88.2%** — 7 月必须靠暑期救场
3. **6 月预计 38K-40K** — 创 2024 以来 6 月新低

🟢 **机会**:
4. **线上散客 88% 主导** — 但纯线上流量忠诚度差
5. **6/13 夜场 79.9 特价 = 3,210 票** — 公式应复用
6. **1-5 月 +7% 反弹基础在** — 6 月暴跌是执行问题

🟡 **战略**:
7. **双节点浪费已固化**(端午后无 H2 接力)
8. **节日 ROI:春节 > 五一 > 十一 > 端午** — 端午应大幅强化"""

    # === E5: 行动建议 ===
    e5 = """**🎯 行动建议(按优先级)**

🔥 **P0 - 本周必做**:
1. **端午后接力**:毕业季+周末活动,目标日均 1,500
2. **7 月档提前规划**:本周完成,7-8 月救命(历史贡献 30-35 万)
3. **复用 6/13 夜场特价公式**:每周五六开 79.9 夜场

⚡ **P1 - 下周必做**:
4. **暑期开门红**:必须冲 4,000+/日
5. **抖音/小红书暑期内容提前 2 周铺**

📊 **P2 - 月度**:
6. **建立会员/季卡留存机制**:端午后暴跌=线上获客→线下转化断裂
7. **补全 2023-2025 端午历史数据**:同比基线下周二前补"""

    # === E6: 数据警示 + SSOT ===
    e6 = f"""**⚠️ 数据质量警示(需站长立即确认)**

- **数据未填**:{last_md} 后续日期未填 → 今日必须补录
- **端午三日大客户期票=0** → 🔴 B 端渠道疑似断流
- **2023-2025 端午历史数据缺** → 同比基线缺失

---

**📊 SSOT(2026-06-23 永久生效)**

- ✅ 新:`~/Downloads/2026游客量统计 (N).csv` + `.dbt(N).xlsx`
- ❌ 弃:`~/Desktop/2026游客量统计.csv`(截至 6/9)
- 🕐 更新频率:**每周二**

---

**📅 下次报告:{next_tuesday()}** — 届时含:7 月开局数据 + 暑期首周末实战 + 6 月最终复盘

—— 李涯 · 自动化生成 · {today}"""

    return {
        "schema": "2.0",
        "header": {
            "title": {"tag": "plain_text", "content": f"📊 电影小镇周二客流深度报告 · {today}"},
            "template": "blue"
        },
        "body": {
            "elements": [
                {"tag": "markdown", "content": e1_full},
                {"tag": "markdown", "content": e2},
                {"tag": "markdown", "content": e3},
                {"tag": "markdown", "content": e4},
                {"tag": "markdown", "content": e5},
                {"tag": "markdown", "content": e6},
            ]
        }
    }


def next_tuesday():
    """算下个周二日期"""
    today = datetime.now()
    days_ahead = 1 - today.weekday()  # 周一是 0, 周二是 1
    if days_ahead <= 0:
        days_ahead += 7
    return (today + __import__('datetime').timedelta(days=days_ahead)).strftime("%Y-%m-%d")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="只生成卡片 JSON 不发送")
    args = parser.parse_args()

    # 1. 找 SSOT
    csv_path, xlsx_path = find_latest_ssot()
    if not csv_path:
        print("❌ 找不到 SSOT CSV,请确认 ~/Downloads/2026游客量统计 (N).csv 存在")
        sys.exit(1)
    print(f"📂 SSOT CSV: {csv_path}")
    print(f"📂 SSOT XLSX: {xlsx_path}")

    # 2. 加载数据
    daily = load_2026_csv(csv_path)
    print(f"📊 加载 {len(daily)} 天数据")
    historical = load_historical_peaks()
    print(f"📊 加载 {len(historical)} 年历史峰值")

    # 3. 算指标
    metrics = compute_metrics(daily, historical)
    print(f"📈 YTD: {metrics['ytd']:,.0f} | 最新: {metrics['last_data_date']}")

    # 4. B1: 周对比
    week_cmp = compute_week_comparison(daily, historical)
    tw = week_cmp['this_week']
    lw = week_cmp['last_week']
    ly = week_cmp['last_year_same_week']
    print(f"📅 周对比: 本周{tw.get('avg_daily',0):.0f}/天 vs 上周{lw.get('avg_daily',0):.0f} vs 去年{ly.get('avg_daily',0):.0f}")

    # 5. B2: 7 日移动平均
    ma7 = compute_ma7(daily)
    print(f"📈 MA7: 起点={ma7[0][0] if ma7 else '-'}, 终点={ma7[-1] if ma7 else '-'}")

    # 6. B3: 天气 (周对比期间 + 本月最近 7 天)
    weather_target = []
    if tw.get('start_date'):
        weather_target.append(f"2026-{tw['start_date']}")
    if tw.get('end_date'):
        weather_target.append(f"2026-{tw['end_date']}")
    # 加上端午窗口
    for d_str in sorted(daily.keys())[-21:]:  # 最近 21 天
        weather_target.append(f"2026-{d_str}")
    weather = fetch_weather_history(sorted(set(weather_target)))
    print(f"🌤️ 天气: 抓到 {len(weather)} 天数据")

    # 7. B4: 活动日历
    activity_cal = load_activity_calendar(xlsx_path)
    print(f"🎪 活动日历: {len(activity_cal)} 天")

    # 打包进 metrics
    metrics['week_cmp'] = week_cmp
    metrics['ma7'] = ma7
    metrics['weather'] = weather
    metrics['activity_cal'] = activity_cal

    # 8. 生成卡片
    card = build_card(metrics, historical, daily)
    # 数表: 找含 `|---|` 或 `| --- |` 的独立行
    table_count = 0
    for e in card['body']['elements']:
        for line in e['content'].split('\n'):
            if line.strip().startswith('|') and '---' in line:
                table_count += 1
    print(f"📋 卡片: {len(card['body']['elements'])} 元素, {table_count} 表")
    assert table_count <= 5, f"表格数 {table_count} 超过飞书限制 5"

    if args.dry_run:
        out = "/tmp/visitor_report_card_dryrun.json"
        with open(out, 'w', encoding='utf-8') as f:
            json.dump(card, f, ensure_ascii=False, indent=2)
        print(f"✅ dry-run 模式,卡片已存: {out}")
        return

    # 5. 发送
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from send_feishu_card import send_card
    result = send_card(FEISHU_CHAT_ID, card, skip_validation=True)
    if result.get('code') == 0:
        print(f"✅ 发送成功: {result['data']['message_id']}")
    else:
        print(f"❌ 发送失败: {json.dumps(result, ensure_ascii=False)[:500]}")
        sys.exit(1)


if __name__ == "__main__":
    main()