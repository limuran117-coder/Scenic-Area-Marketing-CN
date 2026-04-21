#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import numpy as np
import json
import openpyxl
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ============================================================
# DATA LOADING
# ============================================================

# 1. 2026 visitor CSV
df26 = pd.read_csv('/Users/tianjinzhan/Desktop/2026游客量统计.csv')
df26['日期'] = pd.to_datetime(df26['日期'])
df26_sorted = df26.sort_values('日期').reset_index(drop=True)
# cumulative total
df26_sorted['累计'] = df26_sorted['合计'].cumsum()
# filter valid days (actual visitor > 0)
df26_valid = df26_sorted[df26_sorted['合计'] > 0].copy()

# 2. DH data from CSV (2026)
lines_csv = open('/Users/tianjinzhan/Downloads/2026游客量统计 (2).csv').readlines()
# Row 28 (0-index 27): dates, Row29:场次, Row30:库存, Row31:售卖, Row32:上座率
def parse_floats(row_list, skip=2):
    result = []
    for v in row_list[skip:]:
        v = v.strip()
        if v == '' or v == '#DIV/0!':
            result.append(0.0)
        elif v.endswith('%'):
            result.append(float(v[:-1]))
        else:
            try:
                result.append(float(v))
            except:
                result.append(0.0)
    return result

dates_dh26_raw = lines_csv[27].strip().split(',')[2:]
sessions_dh26 = parse_floats(lines_csv[28].split(','), 2)
cap_dh26 = parse_floats(lines_csv[29].split(','), 2)
sold_dh26 = parse_floats(lines_csv[30].split(','), 2)
occ_dh26 = parse_floats(lines_csv[31].split(','), 2)

# Build DH 2026 monthly
months_dh26 = {}
for i in range(len(dates_dh26_raw)):
    d = dates_dh26_raw[i]
    if not d or cap_dh26[i] == 0:
        continue
    mon = int(d.replace('月','').replace('日',''))
    if mon not in months_dh26:
        months_dh26[mon] = {'cap':0,'sold':0,'occ':[]}
    months_dh26[mon]['cap'] += cap_dh26[i]
    months_dh26[mon]['sold'] += sold_dh26[i]
    months_dh26[mon]['occ'].append(occ_dh26[i])

total_cap_26 = sum(months_dh26[m]['cap'] for m in months_dh26)
total_sold_26 = sum(months_dh26[m]['sold'] for m in months_dh26)
avg_occ_26 = total_sold_26 / total_cap_26 * 100

# 3. Excel 2023 DH monthly
wb = openpyxl.load_workbook('/Users/tianjinzhan/Desktop/2023-2025年门票销售及客流统计数据表.xlsx', data_only=True)
ws23 = wb['2023年']
max_col = ws23.max_column

months_dh23 = {}
for c in range(3, max_col+1):
    d = ws23.cell(row=1, column=c).value
    sold = ws23.cell(row=13, column=c).value
    cap = ws23.cell(row=12, column=c).value
    if d and sold is not None:
        m = d.month
        if m not in months_dh23:
            months_dh23[m] = {'cap':0,'sold':0}
        months_dh23[m]['cap'] += int(cap) if cap else 0
        months_dh23[m]['sold'] += int(sold)

total_sold_23 = sum(months_dh23[m]['sold'] for m in months_dh23)
total_cap_23 = sum(months_dh23[m]['cap'] for m in months_dh23)
avg_occ_23 = total_sold_23 / total_cap_23 * 100

# 4. Excel 2025 DH monthly
ws25 = wb['2025年']
max_col25 = ws25.max_column

months_dh25 = {}
for c in range(3, max_col25+1):
    d = ws25.cell(row=1, column=c).value
    sold = ws25.cell(row=13, column=c).value
    cap = ws25.cell(row=12, column=c).value
    if d and sold is not None:
        m = d.month
        if m not in months_dh25:
            months_dh25[m] = {'cap':0,'sold':0}
        months_dh25[m]['cap'] += int(cap) if cap else 0
        months_dh25[m]['sold'] += int(sold)

total_sold_25 = sum(months_dh25[m]['sold'] for m in months_dh25)
total_cap_25 = sum(months_dh25[m]['cap'] for m in months_dh25)
avg_occ_25 = total_sold_25 / total_cap_25 * 100 if total_cap_25 > 0 else 0

# 5. Competitors data
crawl_data = json.load(open('/tmp/crawl_data.json'))
competitors = crawl_data['competitors']

# ============================================================
# KEY COMPUTED VALUES
# ============================================================
ANNUAL = 1320000
days_passed = 120
latest_ytd = df26_sorted['累计'].iloc[-1]
ytd_pct = latest_ytd / ANNUAL * 100
days_pct = days_passed / 365 * 100

ts = df26['散客小计'].sum()
td = df26['渠道小计'].sum()
散客占比 = ts / (ts + td) * 100

# 7-day MA
df26_sorted['MA7'] = df26_sorted['合计'].rolling(7, min_periods=1).mean()

# Monthly visitor summary for 2026
df26_valid['月份'] = df26_valid['日期'].dt.month
monthly_vis = df26_valid.groupby('月份')['合计'].sum()

# Monthly target (按全年132万比例分配)
monthly_target_26 = {}
monthly_days = {1:31, 2:28, 3:31, 4:30, 5:31, 6:30, 7:31, 8:31, 9:30, 10:31, 11:30, 12:31}
for m in range(1, 5):  # through April
    monthly_target_26[m] = ANNUAL * monthly_days[m] / 365

# DH monthly comparison data
# For 2023: months 1-12, 2025: months 1-12, 2026: months 1-4
def get_monthly_occ(months_dict, month):
    if month in months_dict:
        d = months_dict[month]
        if d['cap'] > 0:
            return d['sold'] / d['cap'] * 100
    return 0

# ============================================================
# PLOTLY FIGURES
# ============================================================

PRIMARY = '#1A5276'
ACCENT = '#E67E22'
GREEN = '#27AE60'
RED = '#E74C3C'
GRAY = '#7F8C8D'

FONT = dict(family="Noto Sans SC, PingFang SC, Microsoft YaHei, sans-serif", size=12, color="#333")

def make_fig():
    return go.Figure()

# ---- fig10: 3-year DH comparison (sold, occ, cap) ----
# Use annual totals for 3 years, and monthly breakdown where available
# 2023, 2024, 2026 - annual sold/capacity/occ
# 2024 monthly cap/sold from Excel 2024 sheet
ws24 = wb['2024年']
max_col24 = ws24.max_column
months_dh24 = {}
for c in range(3, max_col24+1):
    d = ws24.cell(row=1, column=c).value
    sold = ws24.cell(row=9, column=c).value  # row 9 = 门票人数合计
    cap = None
    # 2024 DH rows: need to find DH row
    # For 2024, DH rows not present in Excel. Use total 客流 as proxy for sold.
    # But we need DH-specific. Let's estimate: DH sold ~= 客流 * ratio
    # Actually better to use monthly sum from rows 29-32 of the (2) CSV for 2026
    # For 2024, find from some other way
    pass

# For fig10, show annual totals + monthly bar subplots for the 3 years
# Compute annual totals for 2023 and 2026 from available data
# 2023: total sold 251911, cap 333200
# 2026: total sold 127244, cap 190380 (4 months)

# For 2024, extract from the CSV row 1 (2024年参考 门票销售) for total sold
sold_24_total = 0
sold_24_vals = lines_csv[1].strip().split(',')[2:]
for v in sold_24_vals:
    v = v.strip()
    if v and v.isdigit():
        sold_24_total += int(v)

# For 2024 DH capacity - use monthly from Excel (2024 sheet row 11 = 场次, row 12 = 库存, row 13 = 售卖)
months_dh24 = {}
for c in range(3, max_col24+1):
    d = ws24.cell(row=1, column=c).value
    cap = ws24.cell(row=12, column=c).value
    sold = ws24.cell(row=13, column=c).value
    if d and cap is not None:
        m = d.month
        if m not in months_dh24:
            months_dh24[m] = {'cap':0,'sold':0}
        months_dh24[m]['cap'] += int(cap) if cap else 0
        months_dh24[m]['sold'] += int(sold) if sold else 0

total_cap_24 = sum(months_dh24[m]['cap'] for m in months_dh24)
total_sold_24 = sum(months_dh24[m]['sold'] for m in months_dh24)
avg_occ_24 = total_sold_24 / total_cap_24 * 100 if total_cap_24 > 0 else 0

# For fig10: 2023 full year DH, 2024 not available (no DH rows in sheet)
# Use 2023 annual and 2026 Jan-Apr data only
years_fig10 = ['2023年', '2026年(1-4月)']
solds_fig10 = [total_sold_23, total_sold_26]
caps_fig10 = [total_cap_23, total_cap_26]
occs_fig10 = [avg_occ_23, avg_occ_26]

fig10 = make_subplots(rows=1, cols=3,
    subplot_titles=('售卖人次对比', '上座率对比', '转化率对比'),
    horizontal_spacing=0.15)

# col 1: sold bar
fig10.add_trace(go.Bar(x=years_fig10, y=solds_fig10,
    marker_color=[PRIMARY, ACCENT],
    text=[str(int(x)) for x in solds_fig10], textposition='outside', showlegend=False),
    row=1, col=1)
fig10.update_yaxes(title_text='人次', row=1, col=1)

# col 2: occupancy rate bar
fig10.add_trace(go.Bar(x=years_fig10, y=occs_fig10,
    marker_color=[GREEN if o>=70 else ACCENT if o>=55 else RED for o in occs_fig10],
    text=['%.1f%%' % o for o in occs_fig10], textposition='outside', showlegend=False),
    row=1, col=2)
fig10.update_yaxes(title_text='上座率 %', range=[0,100], row=1, col=2)

# col 3: sold/capacity (conversion) bar
conv_fig10 = [s/c*100 if c > 0 else 0 for s,c in zip(solds_fig10, caps_fig10)]
fig10.add_trace(go.Bar(x=years_fig10, y=conv_fig10,
    marker_color=[PRIMARY, ACCENT],
    text=['%.1f%%' % x for x in conv_fig10], textposition='outside', showlegend=False),
    row=1, col=3)
fig10.update_yaxes(title_text='转化率 %', range=[0,100], row=1, col=3)

fig10.update_layout(title_text='《穿越德化街》年度综合对比', font=FONT, height=300)
fig10_html = fig10.to_html(full_html=False, include_plotlyjs='cdn')[7:-6]

# ---- fig7: DH monthly occupancy 2023 vs 2026 ----
fig7 = go.Figure()
months_all = list(range(1, 13))
occ_23_list = [get_monthly_occ(months_dh23, m) for m in months_all]
occ_26_list = [get_monthly_occ(months_dh26, m) for m in months_all]

fig7.add_trace(go.Bar(name='2023年', x=['%d月' % m for m in months_all], y=occ_23_list,
    marker_color=PRIMARY, opacity=0.7))
fig7.add_trace(go.Bar(name='2026年', x=['%d月' % m for m in months_all], y=occ_26_list,
    marker_color=ACCENT, opacity=0.7))
fig7.update_layout(barmode='group', title='《穿越德化街》上座率月度对比 (2023 vs 2026)',
    yaxis_title='上座率 %', font=FONT, height=300)
fig7_html = fig7.to_html(full_html=False, include_plotlyjs='cdn')[7:-6]

# ---- fig8: 2026 monthly occupancy with color coding ----
fig8 = make_subplots(rows=1, cols=2, horizontal_spacing=0.2,
    subplot_titles=('月度上座率', '月度售卖情况'))
months_26_sorted = sorted(months_dh26.keys())
occ_vals = [months_dh26[m]['sold']/months_dh26[m]['cap']*100 for m in months_26_sorted]
colors8 = [GREEN if o >= 70 else ACCENT if o >= 55 else RED for o in occ_vals]
cap_vals = [months_dh26[m]['cap'] for m in months_26_sorted]
sold_vals = [months_dh26[m]['sold'] for m in months_26_sorted]

fig8.add_trace(go.Bar(x=['%d月' % m for m in months_26_sorted], y=occ_vals,
    marker_color=colors8, text=['%.1f%%' % o for o in occ_vals], textposition='outside',
    showlegend=False), row=1, col=1)
fig8.add_hline(y=avg_occ_26, line_dash='dash', line_color=GRAY,
    annotation_text='均值%.1f%%' % avg_occ_26, row=1, col=1)
fig8.update_yaxes(range=[0,100], title_text='上座率 %', row=1, col=1)

fig8.add_trace(go.Bar(x=['%d月' % m for m in months_26_sorted], y=sold_vals,
    marker_color=GREEN, name='售卖', showlegend=True), row=1, col=2)
fig8.add_trace(go.Bar(x=['%d月' % m for m in months_26_sorted], y=[c-s for c,s in zip(cap_vals,sold_vals)],
    marker_color='#ccc', name='空座', showlegend=True), row=1, col=2)
fig8.update_yaxes(title_text='人次', row=1, col=2)
fig8.update_layout(barmode='stack', title='2026年《穿越德化街》月度详情', font=FONT, height=300)
fig8_html = fig8.to_html(full_html=False, include_plotlyjs='cdn')[7:-6]

# ---- fig9: scatter 观演人次 vs 园区客流 ----
# Use daily data: total daily 客流 vs DH daily sold
# DH daily sold from CSV row31
dates_dh26 = dates_dh26_raw
sold_dh26_dict = {}
for i in range(len(dates_dh26_raw)):
    d = dates_dh26_raw[i].replace('日','').replace('月','-')
    sold_dh26_dict[d] = sold_dh26[i]

x9, y9 = [], []
for _, row in df26_sorted.iterrows():
    d_key = row['日期'].strftime('%m-%d')
    if d_key in sold_dh26_dict and sold_dh26_dict[d_key] > 0:
        x9.append(row['合计'])
        y9.append(sold_dh26_dict[d_key])

# fit trendline
if len(x9) > 2:
    z = np.polyfit(x9, y9, 1)
    p = np.poly1d(z)
    x_line = np.linspace(min(x9), max(x9), 50)
    y_line = p(x_line)
    corr = np.corrcoef(x9, y9)[0,1]
    r2 = corr**2
else:
    x_line, y_line, r2 = [], [], 0

fig9 = make_subplots(rows=1, cols=2, horizontal_spacing=0.25,
    subplot_titles=('观演人次 vs 园区客流', '相关性分析'))
fig9.add_trace(go.Scatter(x=x9, y=y9, mode='markers',
    marker=dict(color=PRIMARY, size=8, opacity=0.6), name='每日数据'))
fig9.add_trace(go.Scatter(x=x_line, y=y_line, mode='lines',
    line=dict(color=RED, dash='dot', width=2), name='趋势线'))
fig9.update_xaxes(title_text='园区客流', row=1, col=1)
fig9.update_yaxes(title_text='观演人次', row=1, col=1)

# bar chart of high-occupancy days
if len(x9) > 0:
    x_bin = pd.cut(x9, bins=[0,1000,5000,10000,20000,100000], labels=['<1k','1k-5k','5k-1w','1w-2w','>2w'])
    counts = pd.Series(x9).groupby(x_bin).count()
    fig9.add_trace(go.Bar(x=counts.index.astype(str), y=counts.values,
        marker_color=ACCENT, showlegend=False), row=1, col=2)
    fig9.update_yaxes(title_text='天数', row=1, col=2)

    note = 'R\u00b2 = %.3f' % r2 if r2 else ''
    fig9.add_annotation(x=0.05, y=0.95, text=note, showarrow=False,
        xref='paper', yref='paper', font=dict(size=12, color=GRAY))

fig9.update_layout(title='观演人次与园区客流相关性', font=FONT, height=300, showlegend=False)
fig9_html = fig9.to_html(full_html=False, include_plotlyjs='cdn')[7:-6]

# ---- fig4: donut chart 散客/渠道占比 ----
fig4 = go.Figure()
fig4.add_trace(go.Pie(labels=['散客', '渠道'],
    values=[ts, td],
    hole=0.55,
    marker_colors=[PRIMARY, ACCENT],
    textinfo='label+percent',
    textfont_size=13))
fig4.update_layout(title='散客 vs 渠道占比', font=FONT, height=300,
    annotations=[dict(text='%.1f%%' % 散客占比, x=0.5, y=0.5, font_size=18, showarrow=False)])
fig4_html = fig4.to_html(full_html=False, include_plotlyjs='cdn')[7:-6]

# ---- fig1: daily客流 with 7-day MA ----
fig1 = go.Figure()
fig1.add_trace(go.Scatter(x=df26_sorted['日期'], y=df26_sorted['合计'],
    mode='lines', line=dict(color=PRIMARY, width=1.5), name='每日客流',
    fill='tozeroy', fillcolor='rgba(26,82,118,0.1)'))
fig1.add_trace(go.Scatter(x=df26_sorted['日期'], y=df26_sorted['MA7'],
    mode='lines', line=dict(color=RED, width=2.5), name='7日均线'))
fig1.update_layout(title='2026年每日客流趋势', xaxis_title='日期', yaxis_title='人次',
    font=FONT, height=300)
fig1_html = fig1.to_html(full_html=False, include_plotlyjs='cdn')[7:-6]

# ---- fig2: monthly客流 vs 目标 ----
fig2 = go.Figure()
months_vis = sorted(monthly_vis.index.tolist())
actual_vals = [monthly_vis[m] for m in months_vis]
target_vals = [monthly_target_26.get(m, 0) for m in months_vis]

fig2.add_trace(go.Bar(x=['%d月' % m for m in months_vis], y=actual_vals,
    marker_color=PRIMARY, name='实际客流'))
fig2.add_trace(go.Scatter(x=['%d月' % m for m in months_vis], y=target_vals,
    mode='lines+markers+text', line=dict(color=RED, dash='dot', width=2),
    marker=dict(color=RED, size=8), text=['%d' % int(t) for t in target_vals],
    textposition='top center', name='月度目标'))
fig2.update_layout(title='2026年月度客流 vs 目标', xaxis_title='月份', yaxis_title='人次',
    font=FONT, height=300, barmode='group')
fig2_html = fig2.to_html(full_html=False, include_plotlyjs='cdn')[7:-6]

# ---- fig3: YTD累计进度 ----
fig3 = go.Figure()
fig3.add_trace(go.Scatter(x=df26_sorted['日期'], y=df26_sorted['累计'],
    mode='lines', line=dict(color=PRIMARY, width=2),
    fill='tozeroy', fillcolor='rgba(26,82,118,0.15)', name='累计客流'))
fig3.add_trace(go.Scatter(x=df26_sorted['日期'],
    y=[ANNUAL * (df26_sorted['日期'].iloc[i] - df26_sorted['日期'].iloc[0]).days / 365 for i in range(len(df26_sorted))],
    mode='lines', line=dict(color=RED, dash='dot', width=2), name='时间进度目标线'))
fig3.update_layout(title='YTD累计客流 vs 时间进度', xaxis_title='日期', yaxis_title='累计人次',
    font=FONT, height=300)
fig3_html = fig3.to_html(full_html=False, include_plotlyjs='cdn')[7:-6]

# ---- fig6: 抖音指数 8 competitors ----
comp_names = [c['name'] for c in competitors]
comp_synth = [c['synth'] for c in competitors]
comp_search = [c['search'] for c in competitors]

fig6 = make_subplots(rows=1, cols=2, horizontal_spacing=0.2,
    subplot_titles=('综合指数', '搜索指数'))
fig6.add_trace(go.Bar(y=comp_names, x=comp_synth, orientation='h',
    marker_color=[PRIMARY if n != '建业电影小镇' else ACCENT for n in comp_names],
    text=comp_synth, textposition='outside', showlegend=False), row=1, col=1)
fig6.update_xaxes(title_text='综合指数', row=1, col=1)
fig6.add_trace(go.Bar(y=comp_names, x=comp_search, orientation='h',
    marker_color=[PRIMARY if n != '建业电影小镇' else ACCENT for n in comp_names],
    text=comp_search, textposition='outside', showlegend=False), row=1, col=2)
fig6.update_xaxes(title_text='搜索指数', row=1, col=2)
fig6.update_layout(title='抖音指数竞品对比', font=FONT, height=320)
fig6_html = fig6.to_html(full_html=False, include_plotlyjs='cdn')[7:-6]

# ---- fig5: scatter 天气 vs 客流 ----
# Map weather to numeric impact
weather_map = {
    '正常天': 1, '正常天早行雪化': 0.9, '正常天早上雾': 0.95,
    '阴': 0.8, '阴天': 0.8, '雨': 0.5, '小雨': 0.6, '雨雪天': 0.5,
    '雪': 0.4, '雨夹雪（路面结冰）': 0.3, '大雪闭园': 0,
    '正常天闭园': 0, '大风天': 0.7,
}
w_score = []
w_visitors = []
for _, row in df26_valid.iterrows():
    w = str(row.get('天气', ''))
    score = weather_map.get(w, 0.8)
    w_score.append(score)
    w_visitors.append(row['合计'])

fig5 = go.Figure()
fig5.add_trace(go.Scatter(x=w_score, y=w_visitors, mode='markers',
    marker=dict(color=PRIMARY, size=8, opacity=0.5)))
# trendline
if len(w_score) > 2:
    z5 = np.polyfit(w_score, w_visitors, 1)
    p5 = np.poly1d(z5)
    x5_line = np.linspace(0, 1, 50)
    fig5.add_trace(go.Scatter(x=x5_line, y=p5(x5_line), mode='lines',
        line=dict(color=RED, dash='dot', width=2), showlegend=False))

fig5.update_layout(title='天气影响分析 (天气指数 vs 客流)',
    xaxis_title='天气指数 (0=恶劣, 1=良好)', yaxis_title='日客流',
    font=FONT, height=300)
fig5_html = fig5.to_html(full_html=False, include_plotlyjs='cdn')[7:-6]

# ============================================================
# BUILD HTML
# ============================================================

update_date = '2026年4月21日'
crawl_date = crawl_data.get('date', '2026-04-21')

# KPI values
kpi_ytd = int(latest_ytd)
kpi_pct = ytd_pct
kpi_target_pct = days_pct
kpi_status = '超额' if kpi_pct >= kpi_target_pct else '落后'
kpi_status_color = GREEN if kpi_status == '超额' else RED
kpi_day_avg = int(latest_ytd / days_passed) if days_passed > 0 else 0
kpi_散客 = '%.1f%%' % 散客占比
kpi_dh_occ = '%.1f%%' % avg_occ_26

# Month names for display
month_names = {
    1:'1月', 2:'2月', 3:'3月', 4:'4月', 5:'5月', 6:'6月',
    7:'7月', 8:'8月', 9:'9月', 10:'10月', 11:'11月', 12:'12月'
}

# Section builders
def section(title, chart_html, full_width=False):
    cls = 'chart-wrap' if not full_width else 'chart-wrap full'
    return ('<div class="%s"><div class="section-title">%s</div>'
            '<div class="chart-container">%s</div></div>') % (cls, title, chart_html)

html = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>建业电影小镇 2026 数据看板</title>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@300;400;500;700&display=swap" rel="stylesheet">
<style>
*{margin:0;padding:0;box-sizing:border-box;}
body{font-family:'Noto Sans SC','PingFang SC','Microsoft YaHei',sans-serif;
  background:#F8FAFB;color:#333;font-size:14px;}

/* Header */
.header{background:linear-gradient(135deg,#1A5276 0%,#0E3A5C 100%);
  color:#fff;padding:24px 48px;display:flex;justify-content:space-between;
  align-items:center;}
.header h1{font-size:22px;font-weight:700;letter-spacing:1px;}
.header .subtitle{font-size:13px;opacity:0.8;margin-top:4px;}
.header-right{text-align:right;font-size:13px;opacity:0.85;}

/* KPI row */
.kpi-row{display:grid;grid-template-columns:repeat(5,1fr);gap:16px;
  padding:20px 48px;background:#fff;}
.kpi-card{background:#fff;border-radius:10px;padding:18px 20px;
  box-shadow:0 2px 8px rgba(0,0,0,0.08);text-align:center;border:1px solid #eee;}
.kpi-label{font-size:12px;color:#7F8C8D;text-transform:uppercase;letter-spacing:0.5px;
  margin-bottom:8px;}
.kpi-value{font-size:26px;font-weight:700;color:#1A5276;line-height:1.1;}
.kpi-value .unit{font-size:13px;color:#7F8C8D;font-weight:400;}
.kpi-sub{font-size:11px;color:#aaa;margin-top:4px;}
.badge{display:inline-block;padding:2px 8px;border-radius:12px;font-size:11px;
  font-weight:600;color:#fff;margin-top:4px;}
.badge-green{background:#27AE60;}
.badge-red{background:#E74C3C;}

/* Section */
.section{padding:16px 48px;}
.section-title{font-size:15px;font-weight:600;color:#1A5276;
  margin-bottom:12px;padding-left:4px;
  border-left:3px solid #1A5276;}

/* Chart layout */
.two-col{display:grid;grid-template-columns:1fr 1fr;gap:20px;padding:0 48px;}
.chart-wrap{background:#fff;border-radius:10px;padding:24px;
  box-shadow:0 2px 8px rgba(0,0,0,0.06);margin-bottom:20px;}
.chart-wrap.full{grid-column:1/-1;}
.chart-container{overflow:hidden;}
.section-title{font-size:15px;font-weight:600;color:#1A5276;
  margin-bottom:12px;}

/* Banner */
.banner{background:linear-gradient(135deg,#1A5276 0%,#0E3A5C 100%);
  color:#fff;text-align:center;padding:16px 48px;margin:0 48px;
  border-radius:8px;margin-bottom:20px;}
.banner h2{font-size:16px;font-weight:700;letter-spacing:2px;}
.banner p{font-size:12px;opacity:0.75;margin-top:4px;}

/* Footer */
.footer{text-align:center;padding:24px;color:#aaa;font-size:12px;}

/* Responsive */
@media(max-width:1100px){
  .header{padding:16px 24px;}
  .kpi-row{grid-template-columns:repeat(3,1fr);padding:16px 24px;gap:10px;}
  .two-col,.section{padding:0 24px;}
  .banner{margin:0 24px;}
}
@media(max-width:700px){
  .kpi-row{grid-template-columns:repeat(2,1fr);}
}
</style>
</head>
<body>

<!-- Header -->
<div class="header">
  <div>
    <h1>建业电影小镇 2026 数据看板</h1>
    <div class="subtitle">景区营销中心 · 实时运营监控</div>
  </div>
  <div class="header-right">
    <div>数据更新：%(update_date)s</div>
    <div>抖音抓取：%(crawl_date)s</div>
  </div>
</div>

<!-- KPI Row -->
<div class="kpi-row">
  <div class="kpi-card">
    <div class="kpi-label">累计客流</div>
    <div class="kpi-value">%(kpi_ytd)s<span class="unit"> 人</span></div>
    <div class="kpi-sub">年度目标 132万</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-label">YTD完成率</div>
    <div class="kpi-value">%(kpi_pct).1f<span class="unit">%%</span></div>
    <div class="kpi-sub">时间进度 %(kpi_target_pct).1f%%</div>
    <span class="badge badge-%(kpi_badge)s">%(kpi_status)s</span>
  </div>
  <div class="kpi-card">
    <div class="kpi-label">日均客流</div>
    <div class="kpi-value">%(kpi_day_avg)s<span class="unit"> 人</span></div>
    <div class="kpi-sub">累计 %(days_passed)s 天</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-label">散客占比</div>
    <div class="kpi-value">%(kpi_散客)s</div>
    <div class="kpi-sub">线上+线下+体验</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-label">穿越德化街上座率</div>
    <div class="kpi-value">%(kpi_dh_occ)s</div>
    <div class="kpi-sub">2026累计均值</div>
  </div>
</div>

<!-- DH Banner -->
<div class="section">
<div class="banner">
  <h2>《穿越德化街》专题分析</h2>
  <p>2026年累计观演 %(total_sold_26)s 人次 | 上座率 %(avg_occ_26).1f%% | 总库存 %(total_cap_26)s 人次</p>
</div>
</div>

<!-- Charts row 1: fig10 + fig4 -->
<div class="two-col">
  %(fig10_section)s
  %(fig4_section)s
</div>

<!-- Charts row 2: fig7 + fig8 -->
<div class="two-col">
  %(fig7_section)s
  %(fig8_section)s
</div>

<!-- Charts row 3: fig9 (full) -->
<div class="section">
<div class="chart-wrap full">
  <div class="section-title">观演人次与园区客流相关性分析</div>
  <div class="chart-container">%(fig9)s</div>
</div>
</div>

<!-- Charts row 4: fig1 + fig5 -->
<div class="two-col">
  %(fig1_section)s
  %(fig5_section)s
</div>

<!-- Charts row 5: fig2 + fig3 -->
<div class="two-col">
  %(fig2_section)s
  %(fig3_section)s
</div>

<!-- Charts row 6: fig6 (full) -->
<div class="section">
<div class="chart-wrap full">
  <div class="section-title">抖音指数竞品对比</div>
  <div class="chart-container">%(fig6)s</div>
</div>
</div>

<!-- Footer -->
<div class="footer">
  建业电影小镇 2026 数据看板 · 景区营销中心 · %(footer_date)s
</div>

</body>
</html>""" % {
    'update_date': update_date,
    'crawl_date': crawl_date,
    'kpi_ytd': kpi_ytd,
    'kpi_pct': kpi_pct,
    'kpi_target_pct': kpi_target_pct,
    'kpi_day_avg': kpi_day_avg,
    'kpi_散客': kpi_散客,
    'kpi_dh_occ': kpi_dh_occ,
    'days_passed': days_passed,
    'kpi_status': kpi_status,
    'kpi_badge': 'green' if kpi_status == '超额' else 'red',
    'total_sold_26': int(total_sold_26),
    'total_cap_26': int(total_cap_26),
    'avg_occ_26': avg_occ_26,
    'fig10_section': section('三年《穿越德化街》综合对比', fig10_html),
    'fig4_section': section('散客 vs 渠道占比', fig4_html),
    'fig7_section': section('《穿越德化街》上座率月度对比 (2023 vs 2026)', fig7_html),
    'fig8_section': section('2026年《穿越德化街》月度详情', fig8_html),
    'fig9': fig9_html,
    'fig1_section': section('2026年每日客流趋势', fig1_html),
    'fig5_section': section('天气影响分析', fig5_html),
    'fig2_section': section('2026年月度客流 vs 目标', fig2_html),
    'fig3_section': section('YTD累计客流 vs 时间进度', fig3_html),
    'fig6': fig6_html,
    'footer_date': update_date,
}

out_path = '/Users/tianjinzhan/Desktop/电影小镇2026数据看板.html'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(html)

import os
size = os.path.getsize(out_path)
print('Done! Output: %s' % out_path)
print('File size: %d bytes (%.1f KB)' % (size, size/1024))